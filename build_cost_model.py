"""Build PSP Property Management Cost Projection Model."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

wb = Workbook()
blue_font = Font(name="Arial", color="0000FF", size=11)
black_font = Font(name="Arial", color="000000", size=11)
green_font = Font(name="Arial", color="008000", size=11)
header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
section_font = Font(name="Arial", bold=True, size=12, color="1A3A5C")
title_font = Font(name="Arial", bold=True, size=14, color="1A3A5C")
yellow_fill = PatternFill("solid", fgColor="FFFF00")
header_fill = PatternFill("solid", fgColor="1A3A5C")
light_gray = PatternFill("solid", fgColor="F2F2F2")
gold_fill = PatternFill("solid", fgColor="C4A04D")
green_fill = PatternFill("solid", fgColor="E2EFDA")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
curr = '$#,##0;($#,##0);"-"'
curr2 = '$#,##0.00;($#,##0.00);"-"'
pct = "0.0%"
num = "#,##0"


def hdr(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border


# ── ASSUMPTIONS ──
ws = wb.active
ws.title = "Assumptions"
ws.column_dimensions["A"].width = 42
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 50
ws.merge_cells("A1:C1")
ws["A1"] = "PSP Property Management - Cost Model Assumptions"
ws["A1"].font = title_font

assumptions = [
    (3, "TICKET VOLUME", None, None, True),
    (4, "Tickets/week - Year 1 (beta + rollout)", 30, "Ramp from 1 store to 84", False),
    (5, "Tickets/week - Year 2 (full 84 stores)", 60, "All stores active", False),
    (6, "Tickets/week - Year 3 (100+ stores, new clients)", 100, "Growth from new clients", False),
    (7, "Photos per ticket (avg)", 2, "Compressed to ~350KB", False),
    (8, "Compressed photo size (KB)", 350, "After auto-compression", False),
    (10, "AI USAGE PER TICKET", None, None, True),
    (11, "Warranty lookup input tokens", 2000, "Claude Sonnet", False),
    (12, "Warranty lookup output tokens", 500, "", False),
    (13, "Cost estimation input tokens", 500, "Claude Sonnet", False),
    (14, "Cost estimation output tokens", 200, "", False),
    (15, "Tavily searches per ticket", 3, "Advanced = 2 credits each", False),
    (16, "Tavily credits per ticket", 6, "3 x 2 credits", False),
    (18, "SERVICE PRICING", None, None, True),
    (19, "Claude Sonnet input ($/M tokens)", 3.00, "anthropic.com/pricing", False),
    (20, "Claude Sonnet output ($/M tokens)", 15.00, "", False),
    (21, "Tavily free credits/month", 1000, "tavily.com/pricing", False),
    (22, "Tavily paid plan ($/month)", 30, "10K credits", False),
    (23, "Supabase Pro ($/month)", 25, "supabase.com/pricing", False),
    (24, "Supabase storage overage ($/GB)", 0.021, "", False),
    (25, "Streamlit Cloud ($/month)", 0, "Free tier", False),
    (26, "GitHub ($/month)", 0, "Free private repos", False),
    (28, "VALUE ASSUMPTIONS", None, None, True),
    (29, "Property manager salary", 65000, "", False),
    (30, "Time savings from app", 0.50, "50% of PM time freed up", False),
    (31, "Annual value of time saved", "=B29*B30", "", False),
]

for row, label, val, note, is_section in assumptions:
    if is_section:
        ws.cell(row, 1, label).font = section_font
        continue
    ws.cell(row, 1, label).font = black_font
    if isinstance(val, str) and val.startswith("="):
        ws.cell(row, 2).value = val
        ws.cell(row, 2).font = black_font
    else:
        ws.cell(row, 2, val).font = blue_font
        if isinstance(val, (int, float)) and val >= 25:
            ws.cell(row, 2).fill = yellow_fill
    ws.cell(row, 2).number_format = curr if isinstance(val, (int, float)) and val >= 1000 else num
    if note:
        ws.cell(row, 3, note).font = Font(name="Arial", italic=True, color="666666", size=10)


# ── MONTHLY PROJECTIONS ──
ws2 = wb.create_sheet("Monthly Projections")
hdrs = [
    "Month", "Year", "Tickets/Wk", "Tickets/Mo", "Photos/Mo",
    "Storage Added (GB)", "Cumul Storage (GB)",
    "Claude Input Tok", "Claude Output Tok", "Claude Cost",
    "Tavily Credits", "Tavily Cost",
    "Supabase Cost", "Streamlit", "GitHub",
    "TOTAL MONTHLY", "CUMULATIVE",
]
for c, h in enumerate(hdrs, 1):
    ws2.cell(1, c, h)
    ws2.column_dimensions[get_column_letter(c)].width = 15
ws2.column_dimensions["A"].width = 8
hdr(ws2, 1, len(hdrs))

for m in range(1, 37):
    r = m + 1
    yr = 1 if m <= 12 else (2 if m <= 24 else 3)
    tpw = 30 if yr == 1 else (60 if yr == 2 else 100)
    ws2.cell(r, 1, m)
    ws2.cell(r, 2, yr)
    ws2.cell(r, 3, tpw)
    ws2.cell(r, 4).value = f"=C{r}*4.33"
    ws2.cell(r, 4).number_format = num
    ws2.cell(r, 5).value = f"=D{r}*2"
    ws2.cell(r, 5).number_format = num
    ws2.cell(r, 6).value = f"=E{r}*350/1024/1024"
    ws2.cell(r, 6).number_format = "0.00"
    ws2.cell(r, 7).value = f"=F{r}" if m == 1 else f"=G{r-1}+F{r}"
    ws2.cell(r, 7).number_format = "0.00"
    ws2.cell(r, 8).value = f"=D{r}*(2000+500)"
    ws2.cell(r, 8).number_format = num
    ws2.cell(r, 9).value = f"=D{r}*(500+200)"
    ws2.cell(r, 9).number_format = num
    ws2.cell(r, 10).value = f"=(H{r}/1000000*3)+(I{r}/1000000*15)"
    ws2.cell(r, 10).number_format = curr2
    ws2.cell(r, 11).value = f"=D{r}*6"
    ws2.cell(r, 11).number_format = num
    ws2.cell(r, 12).value = f"=IF(K{r}<=1000,0,30)"
    ws2.cell(r, 12).number_format = curr
    ws2.cell(r, 13).value = f"=IF(G{r}<1,0,25)"
    ws2.cell(r, 13).number_format = curr
    ws2.cell(r, 14, 0)
    ws2.cell(r, 14).number_format = curr
    ws2.cell(r, 15, 0)
    ws2.cell(r, 15).number_format = curr
    ws2.cell(r, 16).value = f"=J{r}+L{r}+M{r}+N{r}+O{r}"
    ws2.cell(r, 16).number_format = curr2
    ws2.cell(r, 16).font = Font(name="Arial", bold=True)
    ws2.cell(r, 17).value = f"=P{r}" if m == 1 else f"=Q{r-1}+P{r}"
    ws2.cell(r, 17).number_format = curr
    ws2.cell(r, 17).font = Font(name="Arial", bold=True)
    if m % 2 == 0:
        for c in range(1, 18):
            ws2.cell(r, c).fill = light_gray


# ── ANNUAL SUMMARY ──
ws3 = wb.create_sheet("Annual Summary")
ws3.merge_cells("A1:F1")
ws3["A1"] = "PSP Property Management - Annual Cost Summary"
ws3["A1"].font = title_font
ws3.column_dimensions["A"].width = 24
for c in range(2, 7):
    ws3.column_dimensions[get_column_letter(c)].width = 16

for c, h in enumerate(["Service", "Year 1", "Year 2", "Year 3", "3-Year Total", "% of Total"], 1):
    ws3.cell(3, c, h)
hdr(ws3, 3, 6)

svcs = [("Claude API", "J"), ("Tavily", "L"), ("Supabase", "M"), ("Streamlit Cloud", "N"), ("GitHub", "O")]
for i, (name, col) in enumerate(svcs):
    r = 4 + i
    ws3.cell(r, 1, name).font = black_font
    ws3.cell(r, 2).value = f"=SUM('Monthly Projections'!{col}2:{col}13)"
    ws3.cell(r, 3).value = f"=SUM('Monthly Projections'!{col}14:{col}25)"
    ws3.cell(r, 4).value = f"=SUM('Monthly Projections'!{col}26:{col}37)"
    ws3.cell(r, 5).value = f"=SUM(B{r}:D{r})"
    ws3.cell(r, 6).value = f"=IF(E9=0,0,E{r}/E9)"
    for c in range(2, 6):
        ws3.cell(r, c).number_format = curr
        ws3.cell(r, c).font = black_font
    ws3.cell(r, 5).font = Font(name="Arial", bold=True)
    ws3.cell(r, 6).number_format = pct
    for c in range(1, 7):
        ws3.cell(r, c).border = thin_border

ws3.cell(9, 1, "TOTAL").font = Font(name="Arial", bold=True, size=12)
ws3.cell(9, 1).fill = gold_fill
for c in range(2, 6):
    ws3.cell(9, c).value = f"=SUM({get_column_letter(c)}4:{get_column_letter(c)}8)"
    ws3.cell(9, c).font = Font(name="Arial", bold=True, size=12)
    ws3.cell(9, c).number_format = curr
    ws3.cell(9, c).fill = gold_fill
    ws3.cell(9, c).border = thin_border
ws3.cell(9, 6).value = 1
ws3.cell(9, 6).number_format = pct
ws3.cell(9, 6).fill = gold_fill
ws3.cell(9, 6).border = thin_border

r = 11
metrics = [
    ("Avg Monthly Cost - Year 1", "=B9/12"),
    ("Avg Monthly Cost - Year 2", "=C9/12"),
    ("Avg Monthly Cost - Year 3", "=D9/12"),
    ("Cost per Ticket - Year 1", "=IF(Assumptions!B4*52=0,0,B9/(Assumptions!B4*52))"),
    ("Cost per Ticket - Year 2", "=IF(Assumptions!B5*52=0,0,C9/(Assumptions!B5*52))"),
    ("Cost per Ticket - Year 3", "=IF(Assumptions!B6*52=0,0,D9/(Assumptions!B6*52))"),
]
for label, formula in metrics:
    ws3.cell(r, 1, label).font = black_font
    ws3.cell(r, 2).value = formula
    ws3.cell(r, 2).font = Font(name="Arial", bold=True)
    ws3.cell(r, 2).number_format = curr2
    r += 1

chart = BarChart()
chart.title = "Annual Cost by Service"
chart.y_axis.title = "Cost ($)"
chart.style = 10
chart.width = 20
chart.height = 12
data_ref = Reference(ws3, min_col=2, min_row=3, max_col=4, max_row=8)
cats = Reference(ws3, min_col=1, min_row=4, max_row=8)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats)
ws3.add_chart(chart, "A19")


# ── BREAK-EVEN ──
ws4 = wb.create_sheet("Break-Even Analysis")
ws4.column_dimensions["A"].width = 42
for c in range(2, 5):
    ws4.column_dimensions[get_column_letter(c)].width = 18
ws4.merge_cells("A1:D1")
ws4["A1"] = "PSP Property Management - Break-Even & ROI"
ws4["A1"].font = title_font

ws4.cell(3, 1, "COSTS WITHOUT THE APP").font = section_font
for c, h in enumerate(["", "Year 1", "Year 2", "Year 3"], 1):
    ws4.cell(4, c, h)
hdr(ws4, 4, 4)

costs_data = [
    (5, "Property Manager salary", [65000, 67000, 69000]),
    (6, "Missed warranty claims (est.)", [5000, 15000, 25000]),
    (7, "Inefficient contractor selection (est.)", [3000, 10000, 15000]),
]
for r, label, vals in costs_data:
    ws4.cell(r, 1, label).font = black_font
    for c, v in enumerate(vals, 2):
        ws4.cell(r, c, v).font = blue_font
        ws4.cell(r, c).fill = yellow_fill
        ws4.cell(r, c).number_format = curr

ws4.cell(8, 1, "Total Hidden Costs").font = Font(name="Arial", bold=True)
for c in range(2, 5):
    cl = get_column_letter(c)
    ws4.cell(8, c).value = f"=SUM({cl}5:{cl}7)"
    ws4.cell(8, c).font = Font(name="Arial", bold=True)
    ws4.cell(8, c).number_format = curr

ws4.cell(10, 1, "APP INFRASTRUCTURE COST").font = section_font
ws4.cell(11, 1, "Annual app cost").font = black_font
ws4.cell(11, 2).value = "='Annual Summary'!B9"
ws4.cell(11, 3).value = "='Annual Summary'!C9"
ws4.cell(11, 4).value = "='Annual Summary'!D9"
for c in range(2, 5):
    ws4.cell(11, c).font = green_font
    ws4.cell(11, c).number_format = curr

ws4.cell(13, 1, "SAVINGS WITH THE APP").font = section_font
for c, h in enumerate(["", "Year 1", "Year 2", "Year 3"], 1):
    ws4.cell(14, c, h)
hdr(ws4, 14, 4)

savings_data = [
    (15, "PM Time Saved (50%)", ["=B5*0.5", "=C5*0.5", "=D5*0.5"]),
    (16, "Warranty Claims Caught", ["=B6", "=C6", "=D6"]),
    (17, "Contractor Optimization", ["=B7", "=C7", "=D7"]),
]
for r, label, formulas in savings_data:
    ws4.cell(r, 1, label).font = black_font
    for c, f in enumerate(formulas, 2):
        ws4.cell(r, c).value = f
        ws4.cell(r, c).number_format = curr

ws4.cell(18, 1, "Total Savings").font = Font(name="Arial", bold=True)
for c in range(2, 5):
    cl = get_column_letter(c)
    ws4.cell(18, c).value = f"=SUM({cl}15:{cl}17)"
    ws4.cell(18, c).font = Font(name="Arial", bold=True, color="008000")
    ws4.cell(18, c).number_format = curr
    ws4.cell(18, c).fill = green_fill

ws4.cell(19, 1, "Less: App Cost").font = black_font
for c in range(2, 5):
    cl = get_column_letter(c)
    ws4.cell(19, c).value = f"=-{cl}11"
    ws4.cell(19, c).font = Font(name="Arial", color="FF0000")
    ws4.cell(19, c).number_format = curr

ws4.cell(20, 1, "NET ANNUAL BENEFIT").font = Font(name="Arial", bold=True, size=13)
for c in range(2, 5):
    cl = get_column_letter(c)
    ws4.cell(20, c).value = f"={cl}18+{cl}19"
    ws4.cell(20, c).font = Font(name="Arial", bold=True, size=13, color="008000")
    ws4.cell(20, c).number_format = curr
    ws4.cell(20, c).fill = gold_fill

ws4.cell(21, 1, "ROI (Savings / Cost)").font = Font(name="Arial", bold=True)
for c in range(2, 5):
    cl = get_column_letter(c)
    ws4.cell(21, c).value = f"=IF({cl}11=0,0,{cl}18/{cl}11)"
    ws4.cell(21, c).font = Font(name="Arial", bold=True, color="008000")
    ws4.cell(21, c).number_format = "0.0x"

ws4.cell(23, 1, "BOTTOM LINE").font = section_font
ws4.cell(24, 1, "Infrastructure costs are minimal (<$1/ticket). The app pays for").font = black_font
ws4.cell(25, 1, "itself through warranty catch, contractor optimization, and PM").font = black_font
ws4.cell(26, 1, "time savings. ROI improves as ticket volume grows.").font = black_font

output = r"C:\Users\BretElliott\Plaza Street Partner Dropbox\Bret Elliott\Ram-Z\Accounting\FZ Fees Process\property-management-app\PSP_Property_Management_Cost_Model.xlsx"
wb.save(output)
print("Saved:", output)
